"""
NOVA — modules/reportes.py
Módulo 9: Generación de reportes en Excel y PDF.
"""
import streamlit as st
import pandas as pd
import io
from datetime import date, timedelta, datetime
from config import seccion_titulo, COLOR
from database import (get_tareas, get_pendientes, get_inventario_items,
                      get_movimientos, get_preventivo_registros, get_preventivo_tareas,
                      get_energia_lecturas, get_config)


# ── EXCEL ──────────────────────────────────────────────────────
def _generar_excel(sb, fecha_ini: date, fecha_fin: date) -> bytes:
    """Genera el reporte Excel con múltiples hojas."""
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                 Border, Side, numbers)

    filtros = {
        "fecha_desde": fecha_ini.isoformat(),
        "fecha_hasta": (fecha_fin + timedelta(days=1)).isoformat(),
    }
    tareas     = get_tareas(sb, filtros)
    pendientes = get_pendientes(sb)
    items_inv  = get_inventario_items(sb)
    movs       = get_movimientos(sb)
    t_prev     = get_preventivo_tareas(sb)
    r_prev     = get_preventivo_registros(sb, fecha_desde=fecha_ini.isoformat(), fecha_hasta=fecha_fin.isoformat())
    energia    = get_energia_lecturas(sb)
    cfg        = get_config(sb)

    wb = openpyxl.Workbook()

    # Estilos
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    alt_fill    = PatternFill("solid", fgColor="1A1F2E")
    border_side = Side(style="thin", color="2D3748")
    thin_border = Border(left=border_side, right=border_side,
                         top=border_side, bottom=border_side)

    def _hoja(nombre: str, columnas: list, filas: list):
        ws = wb.create_sheet(nombre)
        ws.sheet_view.showGridLines = False
        # Cabecera
        for col_idx, c in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=col_idx, value=c)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border
        ws.row_dimensions[1].height = 22
        # Filas
        for r_idx, fila in enumerate(filas, 2):
            fill = alt_fill if r_idx % 2 == 0 else PatternFill()
            for c_idx, val in enumerate(fila, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.fill      = fill
                cell.border    = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        # Autowidth
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        return ws

    # ── HOJA 1: Requerimientos ───────────────────────────────────
    _hoja("Requerimientos",
          ["Fecha","Área","Categoría","Detalle","Técnico","Prioridad","Estado"],
          [[t.get("created_at","")[:10], t.get("lugar","-"), t.get("categoria","-"),
            t.get("detalle","-"), t.get("usuario","-"),
            t.get("prioridad","-"), t.get("estado","-")]
           for t in tareas])

    # ── HOJA 2: Pendientes ───────────────────────────────────────
    pend_rows = []
    for p in pendientes:
        try:
            dias = (date.today() - date.fromisoformat(p.get("created_at","2000-01-01")[:10])).days
        except Exception:
            dias = 0
        area_n = p.get("areas",{}).get("nombre","-") if isinstance(p.get("areas"),dict) else "-"
        pend_rows.append([area_n, p.get("descripcion","-"), p.get("prioridad","-"),
                          p.get("estado","-"), dias, p.get("asignado_a","-")])
    _hoja("Pendientes",
          ["Área","Descripción","Prioridad","Estado","Días Abierto","Responsable"],
          pend_rows)

    # ── HOJA 3: Inventario ───────────────────────────────────────
    _hoja("Inventario",
          ["Ítem","Categoría","Stock Actual","Stock Mínimo","Unidad","Ubicación","Estado"],
          [[i["nombre"], i["categoria"], i["stock_actual"], i["stock_minimo"],
            i["unidad"], i.get("ubicacion","-"),
            "✅ OK" if i["stock_actual"] >= i["stock_minimo"] else "⚠️ Bajo mínimo"]
           for i in items_inv])

    # ── HOJA 4: Mantenimiento Preventivo ─────────────────────────
    prev_rows = []
    for t in t_prev:
        regs_t = [r for r in r_prev if r.get("tarea_id")==t["id"] and r.get("cumplido")]
        pct    = int(len(regs_t) / max(1, 1) * 100)
        obs    = "; ".join(r.get("observaciones","") for r in regs_t if r.get("observaciones"))
        prev_rows.append([t["nombre"], t["frecuencia"].title(), len(regs_t), obs[:200]])
    _hoja("Mantenimiento Preventivo",
          ["Tarea","Frecuencia","Veces Completada (período)","Observaciones"],
          prev_rows)

    # ── HOJA 5: Energía ──────────────────────────────────────────
    _hoja("Energía",
          ["Fecha","Tipo","Lectura","Unidad","Técnico","Notas"],
          [[l["fecha"], l["tipo"].title(), l["valor"], l["unidad"],
            l.get("usuario","-"), l.get("notas","")]
           for l in sorted(energia, key=lambda x: x.get("fecha",""))])

    # ── HOJA 6: Resumen Ejecutivo ─────────────────────────────────
    ws_res = wb.create_sheet("Resumen Ejecutivo")
    ws_res.sheet_view.showGridLines = False
    nombre_hotel = cfg.get("nombre_hotel","Hotel NOVA")
    kpis = [
        ("Hotel",                    nombre_hotel),
        ("Período",                  f"{fecha_ini} → {fecha_fin}"),
        ("Generado el",              date.today().isoformat()),
        ("",""),
        ("Total Requerimientos",     len(tareas)),
        ("Pendientes Abiertos",      sum(1 for p in pendientes if p.get("estado")=="Abierto")),
        ("Pendientes Resueltos",     sum(1 for p in pendientes if p.get("estado")=="Resuelto")),
        ("Items con Stock Bajo",     sum(1 for i in items_inv if i["stock_actual"]<i["stock_minimo"])),
        ("Preventivos Completados",  sum(1 for r in r_prev if r.get("cumplido"))),
    ]
    for row_i, (k, v) in enumerate(kpis, 1):
        ws_res.cell(row=row_i, column=1, value=k).font = Font(bold=True, color="94A3B8", name="Calibri")
        ws_res.cell(row=row_i, column=2, value=v).alignment = Alignment(horizontal="left")
    ws_res.column_dimensions["A"].width = 28
    ws_res.column_dimensions["B"].width = 32

    # Eliminar hoja vacía inicial
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ────────────────────────────────────────────────────────
def _generar_pdf(sb, fecha_ini: date, fecha_fin: date) -> bytes:
    """Genera reporte PDF con ReportLab."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable, PageBreak)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    cfg        = get_config(sb)
    nombre_hotel = cfg.get("nombre_hotel","Hotel NOVA")
    jefe       = cfg.get("jefe_ingenieria","Jefe de Ingeniería")
    filtros    = {"fecha_desde": fecha_ini.isoformat(), "fecha_hasta": (fecha_fin+timedelta(days=1)).isoformat()}
    tareas     = get_tareas(sb, filtros)
    pendientes = get_pendientes(sb)
    items_inv  = get_inventario_items(sb)
    r_prev     = get_preventivo_registros(sb, fecha_desde=fecha_ini.isoformat(), fecha_hasta=fecha_fin.isoformat())

    # Colores
    C_NAVY    = colors.HexColor("#1E3A5F")
    C_BLUE    = colors.HexColor("#4A90E2")
    C_GREEN   = colors.HexColor("#38A169")
    C_RED     = colors.HexColor("#E53E3E")
    C_DARK    = colors.HexColor("#0F1117")
    C_GRAY    = colors.HexColor("#94A3B8")
    C_WHITE   = colors.white
    C_ALT     = colors.HexColor("#1A1F2E")

    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=A4,
                              leftMargin=2*cm, rightMargin=2*cm,
                              topMargin=2*cm, bottomMargin=2.5*cm)
    styles = getSampleStyleSheet()
    story  = []

    titulo_s = ParagraphStyle("titulo",  fontName="Helvetica-Bold",  fontSize=26, textColor=C_WHITE,  spaceAfter=6,  alignment=TA_CENTER)
    sub_s    = ParagraphStyle("sub",     fontName="Helvetica",        fontSize=12, textColor=C_GRAY,   spaceAfter=4,  alignment=TA_CENTER)
    h2_s     = ParagraphStyle("h2",      fontName="Helvetica-Bold",   fontSize=14, textColor=C_BLUE,   spaceAfter=8,  spaceBefore=14)
    body_s   = ParagraphStyle("body",    fontName="Helvetica",         fontSize=9,  textColor=C_WHITE,  leading=14)
    small_s  = ParagraphStyle("small",   fontName="Helvetica",         fontSize=8,  textColor=C_GRAY)

    def _tabla(datos, col_widths, header_row=True):
        t = Table(datos, colWidths=col_widths, repeatRows=1 if header_row else 0)
        style = [
            ("BACKGROUND",  (0,0), (-1,0), C_NAVY),
            ("TEXTCOLOR",   (0,0), (-1,0), C_WHITE),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,0), 8),
            ("ALIGN",       (0,0), (-1,-1), "LEFT"),
            ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
            ("FONTNAME",    (0,1), (-1,-1), "Helvetica"),
            ("FONTSIZE",    (0,1), (-1,-1), 8),
            ("TEXTCOLOR",   (0,1), (-1,-1), C_WHITE),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [C_DARK, C_ALT]),
            ("GRID",        (0,0), (-1,-1), 0.25, colors.HexColor("#2D3748")),
            ("TOPPADDING",  (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0), (-1,-1), 5),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
        ]
        t.setStyle(TableStyle(style))
        return t

    # ── PORTADA ──────────────────────────────────────────────────
    story.append(Spacer(1, 3*cm))
    story.append(Paragraph("⚡ NOVA", titulo_s))
    story.append(Paragraph("Sistema de Gestión de Ingeniería Hotelera", sub_s))
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=C_BLUE))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(f"<b>{nombre_hotel}</b>", ParagraphStyle("hn", fontName="Helvetica-Bold", fontSize=18, textColor=C_WHITE, alignment=TA_CENTER)))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(f"Reporte de Ingeniería · {fecha_ini} al {fecha_fin}", sub_s))
    story.append(Paragraph(f"Generado el {date.today().strftime('%d/%m/%Y')}", small_s))
    story.append(PageBreak())

    # ── RESUMEN EJECUTIVO ─────────────────────────────────────────
    story.append(Paragraph("📊 Resumen Ejecutivo", h2_s))
    pend_abiertos  = sum(1 for p in pendientes if p.get("estado")=="Abierto")
    pend_resueltos = sum(1 for p in pendientes if p.get("estado")=="Resuelto")
    bajo_min       = sum(1 for i in items_inv if i["stock_actual"]<i["stock_minimo"])
    prev_ok        = sum(1 for r in r_prev if r.get("cumplido"))

    kpi_data = [
        ["Indicador", "Valor"],
        ["Total Requerimientos en el período", str(len(tareas))],
        ["Pendientes Abiertos actualmente",    str(pend_abiertos)],
        ["Pendientes Resueltos en el período", str(pend_resueltos)],
        ["Ítems de inventario bajo mínimo",    str(bajo_min)],
        ["Preventivos completados en período", str(prev_ok)],
    ]
    story.append(_tabla(kpi_data, [12*cm, 4*cm]))
    story.append(Spacer(1, 0.8*cm))

    # ── REQUERIMIENTOS ────────────────────────────────────────────
    story.append(Paragraph("📋 Requerimientos del Período", h2_s))
    if tareas:
        req_data = [["Fecha","Área","Categoría","Prioridad","Estado"]]
        for t in tareas[:50]:
            req_data.append([
                t.get("created_at","")[:10], t.get("lugar","-")[:20],
                t.get("categoria","-")[:18], t.get("prioridad","-"), t.get("estado","-"),
            ])
        story.append(_tabla(req_data, [2.5*cm, 4*cm, 3.5*cm, 2.2*cm, 2.8*cm]))
        if len(tareas) > 50:
            story.append(Paragraph(f"… y {len(tareas)-50} requerimientos más. Ver Excel para listado completo.", small_s))
    else:
        story.append(Paragraph("Sin requerimientos en este período.", body_s))
    story.append(Spacer(1, 0.8*cm))

    # ── PENDIENTES CRÍTICOS ───────────────────────────────────────
    story.append(Paragraph("🚨 Pendientes Críticos", h2_s))
    criticos = [p for p in pendientes if p.get("prioridad")=="Alta" and p.get("estado")!="Resuelto"]
    if criticos:
        pend_data = [["Área","Descripción","Días Abierto","Responsable"]]
        for p in criticos:
            try:
                dias = (date.today() - date.fromisoformat(p.get("created_at","2000-01-01")[:10])).days
            except Exception:
                dias = 0
            area_n = p.get("areas",{}).get("nombre","-") if isinstance(p.get("areas"),dict) else "-"
            pend_data.append([area_n[:20], p.get("descripcion","")[:40], str(dias), p.get("asignado_a","-")[:18]])
        story.append(_tabla(pend_data, [3.5*cm, 6*cm, 2.5*cm, 3*cm]))
    else:
        story.append(Paragraph("✅ Sin pendientes críticos actualmente.", body_s))
    story.append(Spacer(1, 0.8*cm))

    # ── INVENTARIO CRÍTICO ────────────────────────────────────────
    story.append(Paragraph("📦 Inventario Bajo Mínimo", h2_s))
    criticos_inv = [i for i in items_inv if i["stock_actual"] < i["stock_minimo"]]
    if criticos_inv:
        inv_data = [["Ítem","Categoría","Stock Actual","Stock Mínimo","Unidad"]]
        for i in criticos_inv:
            inv_data.append([i["nombre"][:30], i["categoria"], str(i["stock_actual"]),
                             str(i["stock_minimo"]), i["unidad"]])
        story.append(_tabla(inv_data, [5*cm, 3.5*cm, 2.5*cm, 2.5*cm, 2*cm]))
    else:
        story.append(Paragraph("✅ Todo el inventario está sobre el mínimo.", body_s))

    # ── PIE DE PÁGINA / FIRMA ─────────────────────────────────────
    story.append(Spacer(1, 2*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_GRAY))
    story.append(Spacer(1, 0.3*cm))
    firma_data = [
        [Paragraph(f"<b>{jefe}</b>", ParagraphStyle("f1",fontName="Helvetica-Bold",fontSize=10,textColor=C_WHITE)),
         Paragraph(f"<b>Generado por NOVA</b>", ParagraphStyle("f2",fontName="Helvetica-Bold",fontSize=10,textColor=C_WHITE,alignment=TA_CENTER)),
         Paragraph(f"<b>{date.today().strftime('%d/%m/%Y')}</b>", ParagraphStyle("f3",fontName="Helvetica-Bold",fontSize=10,textColor=C_WHITE))],
        [Paragraph("Firma del responsable",small_s), Paragraph(nombre_hotel,small_s), Paragraph("Fecha",small_s)],
    ]
    t_firma = Table(firma_data, colWidths=[5*cm, 6*cm, 5*cm])
    t_firma.setStyle(TableStyle([("ALIGN",(0,0),(-1,-1),"LEFT"), ("VALIGN",(0,0),(-1,-1),"TOP")]))
    story.append(t_firma)

    doc.build(story)
    return buf.getvalue()


# ── RENDER DEL MÓDULO ──────────────────────────────────────────
def render(sb):
    seccion_titulo("📄", "Reportes", "Exportación de datos en Excel y PDF")

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">⚙️ Configurar Reporte</div>', unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    periodo = c1.selectbox("Período predefinido",
                           ["Hoy","Esta semana","Este mes","Rango personalizado"],
                           key="rep_per")
    hoy = date.today()
    if periodo == "Hoy":
        fi, ff = hoy, hoy
    elif periodo == "Esta semana":
        fi, ff = hoy - timedelta(days=hoy.weekday()), hoy
    elif periodo == "Este mes":
        fi, ff = hoy.replace(day=1), hoy
    else:
        fi = c2.date_input("Desde", value=hoy.replace(day=1), key="rep_fi")
        ff = c2.date_input("Hasta", value=hoy, key="rep_ff")

    st.markdown(f"""
    <div style="background:rgba(74,144,226,0.1);border:1px solid rgba(74,144,226,0.3);
                border-radius:10px;padding:12px 16px;margin:12px 0;font-size:0.88rem;color:#90cdf4;">
        📅 Período seleccionado: <strong>{fi}</strong> → <strong>{ff}</strong>
    </div>
    """, unsafe_allow_html=True)

    c_excel, c_pdf = st.columns(2)

    with c_excel:
        st.markdown('<div class="btn-success">', unsafe_allow_html=True)
        if st.button("📊 Generar Excel", use_container_width=True, key="btn_excel"):
            with st.spinner("Generando Excel..."):
                try:
                    data_excel = _generar_excel(sb, fi, ff)
                    st.download_button(
                        label="⬇️ Descargar Excel",
                        data=data_excel,
                        file_name=f"NOVA_Reporte_{fi}_{ff}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_excel",
                    )
                    st.success("✅ Excel generado.")
                except Exception as e:
                    st.error(f"❌ Error generando Excel: {e}")
        st.markdown('</div>', unsafe_allow_html=True)

    with c_pdf:
        st.markdown('<div class="btn-warning">', unsafe_allow_html=True)
        if st.button("📄 Generar PDF", use_container_width=True, key="btn_pdf"):
            with st.spinner("Generando PDF..."):
                try:
                    data_pdf = _generar_pdf(sb, fi, ff)
                    st.download_button(
                        label="⬇️ Descargar PDF",
                        data=data_pdf,
                        file_name=f"NOVA_Reporte_{fi}_{ff}.pdf",
                        mime="application/pdf",
                        key="dl_pdf",
                    )
                    st.success("✅ PDF generado.")
                except Exception as e:
                    st.error(f"❌ Error generando PDF: {e}")
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

    # Info
    st.markdown("""
    <div class="card">
        <div class="card-title">📋 Contenido de los reportes</div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;font-size:0.85rem;color:#94a3b8;">
            <div>
                <strong style="color:#68d391;">📊 Excel incluye:</strong><br>
                • Hoja 1: Requerimientos del período<br>
                • Hoja 2: Pendientes y estados<br>
                • Hoja 3: Inventario completo<br>
                • Hoja 4: Mantenimiento preventivo<br>
                • Hoja 5: Lecturas de energía<br>
                • Hoja 6: Resumen ejecutivo con KPIs
            </div>
            <div>
                <strong style="color:#90cdf4;">📄 PDF incluye:</strong><br>
                • Portada con nombre del hotel<br>
                • Resumen ejecutivo con KPIs<br>
                • Tabla de requerimientos<br>
                • Pendientes críticos<br>
                • Inventario bajo mínimo<br>
                • Firma del jefe de ingeniería
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
